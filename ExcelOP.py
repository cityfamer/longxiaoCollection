import openpyxl
import os


class ExcelOP(object):
    def __init__(self):
        filename = "./data/利能项目_20210916222206.xlsx"
        wb1 = openpyxl.load_workbook(filename,data_only=True)
        self.linengsheet = wb1.get_sheet_by_name("Worksheet")

        filename = "./data/非利能项目_20210916152610.xlsx"
        wb2 = openpyxl.load_workbook(filename,data_only=True)

        self.nolinengsheet = wb2.get_sheet_by_name("Worksheet")

        filename = "./data/收款明细-重庆及挂靠2016-2021（总）.xlsx"
        wb3 = openpyxl.load_workbook(filename,data_only=True)

        self.havegetmoneysheet = wb3.get_sheet_by_name("2016-2018重庆")

        self.totalprojectdata = []
        self.havegetmoneyprojectdata = []

    def save(self):
        self.wb.save(self.filename)

    def getprojectdata(self):
        self.getlinengprojectdata()
        self.getnolinengprojectdata()
        self.gethavegetmoneyprojectdata()
        return self.totalprojectdata,self.havegetmoneyprojectdata

    def getlinengprojectdata(self):
        for i in range(2, 2464):
            place = self.linengsheet["B" + str(i)].value
            if (place == "重庆"):
                num = self.linengsheet["G" + str(i)].value
                if num==None: num=""
                num =self.formatnum(num)
                name = self.linengsheet["H" + str(i)].value
                camp="利能"
                getmoney = self.linengsheet["BA" + str(i)].value
                if getmoney!=None:
                    if float(getmoney) > 0:
                        self.totalprojectdata.append([num, name,camp, getmoney])

    def getnolinengprojectdata(self):
        for i in range(2, 3571):
            place = self.nolinengsheet["C" + str(i)].value
            if (place == "重庆"):
                num = self.nolinengsheet["H" + str(i)].value
                if num==None: num=""
                num =self.formatnum(num)
                name = self.nolinengsheet["I" + str(i)].value
                camp= self.nolinengsheet["K" + str(i)].value
                getmoney = self.nolinengsheet["AS" + str(i)].value
                if getmoney!=None:
                    if float(getmoney) > 0:
                        self.totalprojectdata.append([num, name,camp, getmoney])

    def gethavegetmoneyprojectdata(self):
        for i in range(3, 833):
            num = self.havegetmoneysheet["B" + str(i)].value
            if num==None: num=""
            num =self.formatnum(num)
            name = self.havegetmoneysheet["C" + str(i)].value
            getmoney = self.havegetmoneysheet["I" + str(i)].value
            if getmoney!=None:
                if float(getmoney) >0:
                    self.havegetmoneyprojectdata.append([num, name, getmoney])

    def formatnum(self,num):
        temp_num = num.replace("－", "-")
        temp_num = temp_num.replace(" -", "-")
        temp_num = temp_num.replace("- ", "-")
        temp_num = temp_num.replace("  ", " ")
        temp_num = temp_num.replace("  ", " ")
        temp_num = temp_num.replace("，", "&")
        temp_num = temp_num.replace(",", "&")
        temp_num = temp_num.replace("/", "&")
        temp_num = temp_num.replace(";", "&")
        temp_num = temp_num.replace("、", "&")
        temp_num = temp_num.replace(" ", "&")
        temp_num = temp_num.replace(chr(10), '&')
        temp_num = temp_num.replace("&&", "&");
        temp_num = temp_num.replace("&&", "&");
        return temp_num;
